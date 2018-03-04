from django.shortcuts import render, get_object_or_404, reverse
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from registrations.models import *
# from ems.models import *
from events.models import *
from shop.models import *
from django.contrib.auth.models import User
from .serializers import *
from oasis2017.keyconfig import *
from oasis2017.settings import BASE_DIR
import os
import re
import requests
#sendgrid mailing
import sendgrid
from sendgrid.helpers.mail import *
#DRF
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAuthenticated, BasePermission
from rest_framework.authentication import BasicAuthentication, SessionAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
#instamojo
from instamojo_wrapper import Instamojo
try:
    api = Instamojo(api_key=INSTA_API_KEY, auth_token=AUTH_TOKEN)
except:
    api = Instamojo(api_key=INSTA_API_KEY_test, auth_token=AUTH_TOKEN_test, endpoint='https://test.instamojo.com/api/1.1/') #when in development
#random passwords
import string
from random import choice
chars = string.letters + string.digits
for i in '0oO1QlLiI':
    chars = chars.replace(i,'')
import random
import os
USER_NAMES = ['f20170636', 'f2015831', 'f20170216', 'f2016023', 'f20170647']
#fireabse
import pyrebase
config = {
    "apiKey": FIREBASE_API_KEY,
    "authDomain": "smart-campus-ba95e.firebaseio.com",
	"databaseURL": "https://smart-campus-ba95e.firebaseio.com",
    "storageBucket": "smart-campus-ba95e.appspot.com",
	"serviceAccount": os.path.join(BASE_DIR, "oasis2017/smart-campus-ba95e-firebase-adminsdk-4dl5v-d645a5dd47.json")
}
firebase = pyrebase.initialize_app(config)
db = firebase.database()

#google oauth client side
from google.oauth2 import id_token
from google.auth.transport import requests as requests1

class UserPermissionClass(BasePermission):

	def has_permission(self, request, view):
		user = request.user
		try:
			part = user.participant
			return True
		except:
			try:
				b = Bitsian.objects.get(email=user.email)
				print 2
				return True
			except:
				print 3
				return False

class TokenVerificationClass(BasePermission):

	def has_permission(self, request, view):
		return request.data.get('WALLET_TOKEN') == WALLET_TOKEN

#when a bitsian login through oauth
@api_view(['POST',])
@permission_classes((AllowAny, ))
def get_bt(request):
	token = request.data['id_token']
	idinfo = id_token.verify_oauth2_token(token, requests1.Request(), OAUTH_CLIENT_ID_app)
	print 1
	if idinfo['iss'] not in ['accounts.google.com', 'https://accounts.google.com']:
		return Response({'status':0, 'message':'Invalid user'})
	email = idinfo['email']
	print 2
	try:
		print Bitsian
		bitsian = Bitsian.objects.get(email=email)
		print 3
	except:
		return Response({'message':'Bitsian not found.'})
	password = ''.join(choice(chars) for _ in xrange(8))
	username = email.split('@')[0]
	try:
		user = User.objects.get(username=username)
		user.email = email
		user.set_password(password)
		user.save()
	except:
		user = User.objects.create_user(username=username, password=password, email=email)
		bitsian.user = user
		bitsian.save()
	return Response({'username':username, 'password':password})


@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def get_profile(request):
	try:
		participant = Participant.objects.get(user=request.user)
		if not participant.firewallz_passed:
			return Response({'message':'Register at Firewallz first.'})
		event_set = [participation.event for participation in Participation.objects.filter(participant=participant, pcr_approved=True)]
		event_serializer = EventSerializer(event_set, many=True)
		is_bitsian = False
	except:
		bitsian = Bitsian.objects.get(email=request.user.email)
		is_bitsian = True
	if is_bitsian:
		try:
			wallet = Wallet.objects.get(bitsian=bitsian)
		except:
			wallet = create_wallet(request.user, 9898989898, is_bitsian)
	else:
		try:
			wallet = Wallet.objects.get(participant=participant)
		except:
			wallet = create_wallet(request.user, participant.phone, is_bitsian)
	wallet_serializer = WalletSerializer(wallet)
	list_={}
	for t in wallet.transactions.all():
		list_[t.id] = TransactionSerializer(t).data
	dict_ = {'wallet':wallet_serializer.data, 'transactions': list_}
	ts = TransactionSerializer(wallet.transactions.all(), many=True).data
	db.child('wallet').child(wallet.uid).set(dict_)
	if is_bitsian:
		bitsian_serializer = BitsianSerializer(bitsian)
		# profshow_serializer = AttendanceSerializer(Attendance.objects.filter(bitsian=bitsian), many=True)
		return Response({'bitsian':bitsian_serializer.data, 'wallet':wallet_serializer.data, 'transactions':ts})
	else:
		participant_serializer = ProfileSerializer(participant, context={'request':request})
		# profshow_serializer = AttendanceSerializer(Attendance.objects.filter(participant=participant), many=True)
		return Response({'participant':participant_serializer.data, 'participations':event_serializer.data, 'wallet':wallet_serializer.data, 'transactions':ts})

def create_wallet(user, phone, is_bitsian):
	if is_bitsian:
		bitsian = Bitsian.objects.get(email=user.email)
		wallet, created = Wallet.objects.get_or_create(user=user, bitsian=bitsian, is_bitsian=True)
	else:
		part = user.participant
		wallet, created = Wallet.objects.get_or_create(user=user, participant=part)
	if created:
		wallet.phone = phone
		wallet.save()
	try:
		if not wallet.userid == '':
			pass
		else:
			raise Exception
	except:
		if is_bitsian:
			wallet.userid = bitsian.email.split('@')[0]
		else:
			wallet.userid = str(part.college.id).rjust(4, '0')[1:] + str(part.id).rjust(3, '0')
		wallet.save()
	return wallet


@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def add_money_bitsian(request):		###from swd account
	# user = request.user
	# username = user.username
	# if username not in USER_NAMES:
	# return Response({'status':1, 'message':'Payment portal under construction.'})
	user = request.user
	try:
		wallet = user.wallet
	except:
		return Response({'status':0, 'message':'Wallet does not exist'})
	try:
		amount = int(request.data['amount'])
		if amount <= 0:
			raise Exception
		if amount >= 10000:
			return Response({'status':3, 'message':'Sorry dude! Not allowed'})
	except:
		return Response({'status':2, 'message':'Invalid amount'})
	if not wallet.is_bitsian:
		return Response({'status':0, 'message':'Not Allowed'})

	wallet.curr_balance += amount
	wallet.save()
	transaction = Transaction.objects.create(wallet=wallet, value=amount, t_type="swd")
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)
	return Response({'status':1, 'message':str(amount) + ' Successfully added'})

@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def add_money_request(request):
	return Response({'status':10, 'url':'Payment portal under construction.', 'message':'Payment portal under construction.'})
	user = request.user
	try:
		wallet = Wallet.objects.get(user=user)
	except:
		return Response({'status':3, 'message':'Wallet does not exist.'})
	data = request.data
	try:
		amount = int(data['amount'])
		if amount <= 0:
			raise Exception
	except:
		return Response({'status':0, 'message': 'Enter a valid amount'})
	purpose = 'Add Money to the Apogee wallet'
	phone = int(wallet.phone)
	try:
		part = user.participant
		name = part.name
		email = part.email
	except:
		email = user.email
		bitsian = Bitsian.objects.get(email=email)
		name = user.get_full_name()

	redirect_url = request.build_absolute_uri(reverse("shop:add_money_response"))
	response = api.payment_request_create(buyer_name=name, email=email, amount=amount, purpose=purpose, phone=phone, redirect_url=redirect_url)
	try:
		url = response['payment_request']['longurl']
		return Response({'status':1, 'url':url})
	except:
		return Response({'status':2, 'message':'Payment could not be initialised'})

def add_money_response(request):
	payid=str(request.GET['payment_request_id'])
	try:
		headers = {'X-Api-Key': INSTA_API_KEY, 'X-Auth-Token': AUTH_TOKEN}
   		r = requests.get('https://www.instamojo.com/api/1.1/payment-requests/'+str(payid),headers=headers)
	except:
		headers = {'X-Api-Key': INSTA_API_KEY_test, 'X-Auth-Token': AUTH_TOKEN_test}
		r = requests.get('https://test.instamojo.com/api/1.1/payment-requests/'+str(payid), headers=headers)    ### when in development

	json_ob = r.json()
	if (json_ob['success']):
		payment_request = json_ob['payment_request']
		purpose = payment_request['purpose']
		amount = int(float(payment_request['amount']))
		email = payment_request['email']
		try:
			bitsian = Bitsian.objects.get(email=email)
			wallet = bitsian.wallet
		except:
			part = Participant.objects.get(email=email)
			wallet = part.wallet
		payment_id = payment_request['payments'][0]['payment_id']
		try:
			transaction = Transaction.objects.get(payment_refund_id=payment_id,)
			return HttpResponse('Payment Unsuccessful')
		except:
			pass
		transaction = Transaction.objects.create(value=amount, wallet=wallet, payment_refund_id=payment_id, t_type="add")
		wallet.curr_balance += amount
		wallet.save()
		db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
		db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)
		return HttpResponse('Payment successful')
	return HttpResponse('Payment Unsuccessful')

@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def scan_qr_user(request):
	user = request.user
	try:
		uid = request.data['code']
		wallet = Wallet.objects.get(uid=uid)
	except:
		return Response({'status':0, 'message':'Invalid code'})
	phone = wallet.phone
	balance = wallet.curr_balance
	if wallet.is_bitsian:
		name = wallet.user.get_full_name()
		email = wallet.user.email
	else:
		name = wallet.participant.name
		email = wallet.participant.email
	return Response({'data':{'balance':balance, 'name':name, 'email':email, 'code':code, 'phone':phone}, 'status':1})

@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def transfer_money(request):
	# user = request.user
	# username = user.username
	# if username not in USER_NAMES:
	# return Response({'status':10, 'url':'Payment portal under construction.', 'message':'Payment portal under construction.'})
	data = request.data
	user = request.user
	try:
		wallet = user.wallet
	except:
		return Response({'status':0, 'message':'Wallet does not exist for this user'})
	try:
		uid = request.data['uid']
		wallet_rec = Wallet.objects.get(uid=uid)
		user_rec = wallet_rec.user
	except:
		try:
			wallet_rec = Wallet.objects.get(userid=uid)
			user_rec = wallet_rec.user
		except:
			return Response({'status':3, 'message':'Invalid Code for the user'})
	try:
		amount = int(request.data['amount'])
		if amount<=0:
			raise Exception
	except:
		return Response({'status':2, 'message':'Invalid Amount'})
	if amount > wallet.curr_balance:
		return Response({'status':4, 'message':'Insufficient Balance'})
	if wallet == wallet_rec:
		return Response({'status':5, 'message':'Dude Seriously?'})
	wallet.curr_balance = wallet.curr_balance - amount
	wallet.save()
	wallet_rec.curr_balance = wallet_rec.curr_balance + amount
	wallet_rec.save()
	transaction = Transaction.objects.create(wallet=wallet, value = -amount, t_type="transfer", transfer_to_from=wallet_rec)
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)
	transaction_rec = Transaction.objects.create(wallet=wallet_rec, value = amount, t_type="recieve", transfer_to_from=wallet)
	db.child('wallet').child(wallet_rec.uid).child('transactions').child(transaction_rec.id).set(TransactionSerializer(transaction_rec).data)
	db.child('wallet').child(wallet_rec.uid).child('wallet').set(WalletSerializer(wallet_rec).data)
	return Response({'status':1, 'message':'Money transferred successfully'})

@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass))
def get_stalls(request):
	stalls = Stall.objects.all()
	data = []
	for stall in stalls:
		data.append({'name':stall.name, 'description':stall.description, 'id':stall.id})
	return Response({'stalls':data})

@api_view(['POST',])
@permission_classes((AllowAny,))
def get_products(request, stall_id):
	stall = get_object_or_404(Stall,id=stall_id)
	stall_serializer = StallSerializer(stall)
	products = ProductMainSerializer(ProductMain.objects.filter(product__stall=stall).order_by('-orderno', 'product__name'), many=True)
	return Response({'stalls':stall_serializer.data, 'products':products.data})


@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass, UserPermissionClass))
def checkout_payment(request):
	# user = request.user
	# username = user.username
	# if username not in USER_NAMES:
	# return Response({'status':10, 'url':'Payment portal under construction.', 'message':'Payment portal under construction.'})
	try:
		user = request.user
		wallet = user.wallet
		curr_balance = wallet.curr_balance
	except:
		return Response({'status':0, 'message':'Wallet does not exist'})
	stall_set = []
	try:
		orders = dict(request.data)['current_cart']
		total = 0
		products = []
		for o in orders:
			p = ProductMain.objects.get(id=o['id'])
			if not p.is_available:
				return Response({'status':4, 'message':p.__unicode__() + ' is not available right now.'})
			q = int(o['quantity'])
			products.append([p,q])
			total += int(p.price * q * (100-p.discount)/100)
			stall_set.append(p.product.stall)
	except:
		return Response({'status':3, 'message':'invalid codes for products'})
	from django.db import transaction
	with transaction.atomic():
		wallet = Wallet.objects.select_for_update().get(user=user)
		curr_balance = wallet.curr_balance
		if curr_balance < total:
			return Response({'status':2, 'message':'Insufficient Balance, need to add money to the wallet.'})
		wallet.curr_balance -= total
		stall_set = set(stall_set)
		data = []
		stg_ids = []
		try:
			bitsian = Bitsian.objects.get(email=request.user.email)
			is_bitsian = True
		except:
			is_bitsian=False
			participant = Participant.objects.get(user=request.user)
		for stall in stall_set:
			transaction = Transaction.objects.create(wallet=wallet, t_type="buy")
			if is_bitsian:
				stall_group = StallGroup.objects.create(transaction=transaction, is_bitsian=is_bitsian,stall=stall, user=user, bitsian=bitsian)
			else:
				stall_group = StallGroup.objects.create(transaction=transaction, is_bitsian=is_bitsian,stall=stall, user=user, participant=participant)
			data.append([stall, stall_group, transaction])
			stg_ids.append(stall_group.id)
		for i in products:
			quantity = i[1]
			product = i[0]
			if product.product.p_type.name == 'Apparel':
				product.quantity_left -= quantity
				product.save()
			elif product.product.p_type.name == 'ProfShow':
				prof_show = product.product.prof_show
				s = get_attendance_count(prof_show)
				if is_bitsian:
					try:
						attendance = Attendance.objects.get(bitsian=bitsian, prof_show=prof_show)
						attendance.count += quantity
						attendance.save()
					except:
						attendance = Attendance()
						attendance.bitsian = bitsian
						attendance.prof_show = prof_show
						attendance.paid = True
						attendance.count = quantity
						attendance.save()
				else:
					try:
						attendance = Attendance.objects.get(participant=participant, prof_show=prof_show)
						attendance.count += quantity
						attendance.save()
					except:
						attendance = Attendance()
						attendance.participant = participant
						attendance.prof_show = prof_show
						attendance.paid = True
						attendance.count = quantity
						attendance.save()
				attendance.number = s
				attendance.save()
			stall = product.product.stall
			stall_group = StallGroup.objects.filter(id__in=stg_ids).get(stall=stall)
			transaction = stall_group.transaction
			stall_group.amount += int(product.price * quantity*(100-product.discount)/100)
			stall_group.unique_code = str(random.randint(1000, 9999))
			stall_group.orderid = str(stall_group.id) + stall.name[:3] + wallet.userid
			stall_group.save()
			stall_group.order_no = 1000000 + stall_group.id
			stall_group.save()
			transaction.value += int(product.price * quantity*(100-product.discount)/100)
			transaction.save()
			sale = Sale.objects.create(product=product, quantity=quantity, stall_group=stall_group, paid=True)
			sale.save()
		##for vendor
		for stg_id in stg_ids:
			stg = StallGroup.objects.get(id=stg_id)
			if stg.stall.id == 12:
				stg.code_requested=True
				stg.order_ready=True
				stg.order_complete=True
				stg.save()
			t = stg.transaction
			db.child('stall').child(stg.stall.id).child(stg.id).set(StallGroupSerializer(stg).data)
			db.child('wallet').child(wallet.uid).child('transactions').child(t.id).set(TransactionSerializer(t).data)
	
	##for the user
		wallet.save()
		db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)
		return Response({'status':1, 'message': 'Your order has been placed. Wait for your turn.'})

@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass))
def change_availability(request):
	try:
		user = request.user
		stall = user.stall
	except:
		logout(request)
		return Response({'status':0, 'message':'Invalid User'})
	if request.method == 'POST':
		data = request.data
		try:
			product_id = data['p_id']
			product = ProductMain.objects.get(id=product_id, product__stall=stall)
		except:
			return Response({'status':0, 'message':'Invalid code'})
		if product.is_available:
			product.is_available = False
		else:
			product.is_available = True
		product.save()
	stall_serializer = StallSerializer(stall)
	return JsonResponse({'status':1, 'stall':stall_serializer.data})


@api_view(['POST','OPTIONS', 'GET'])
@permission_classes((IsAuthenticated,))
def sales_today(request):
	import datetime
	today = datetime.datetime.now()
	one_day = datetime.timedelta(days=1)

	today = today.replace(hour=16)
	yesterday = today - one_day
	if today.hour < 3:
		today = today - one_day
		yesterday = yesterday - one_day
	print today
	print yesterday
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	today_sales = SaleSerializer(Sale.objects.filter(stall_group__cancelled=False, product__product__stall=stall, stall_group__transaction__created_at__gte=today), many=True).data
	yesterday_sales = SaleSerializer(Sale.objects.filter(stall_group__cancelled=False, product__product__stall=stall, stall_group__transaction__created_at__gte=yesterday, stall_group__transaction__created_at__lte=today), many=True).data
	return Response({'status':1, 'today':today_sales, 'yesterday':yesterday_sales})

@api_view(['POST','OPTIONS', 'GET'])
@permission_classes((IsAuthenticated,))
def all_sales(request):
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	sales = SaleSerializer(Sale.objects.filter(stall_group__cancelled=False, product__product__stall=stall), many=True).data
	return Response({'status':1, 'sales':sales})


##for laddha
@api_view(['POST','OPTIONS'])
@permission_classes((IsAuthenticated,))
def get_stall_id(request):
	try:
		stall = request.user.stall
		is_open =any([p.is_available for p in ProductMain.objects.filter(product__stall=stall)])
		return Response({'status':1, 'id':stall.id, 'name':stall.name, 'open':is_open})
	except:
		return Response({'status':0})


#cancel laddha's side
@api_view(['POST'])
@permission_classes((IsAuthenticated, ))
def cancel_order(request):
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})

	try:
		sg_id = request.data['sg_id']
		stg = StallGroup.objects.get(id=sg_id)
		transaction = stg.transaction
		wallet = transaction.wallet
		stall = stg.stall
		if stg.order_complete or stg.order_ready:
			raise Exception
	except:
		return Response({'status':2, 'message':'invalid code'})

	total = stg.amount
	stg.sales.all().update(cancelled=True)
	stg.cancelled = True
	stg.save()
	db.child('stall').child(stall.id).child(stg.id).set(StallGroupSerializer(stg).data)
	wallet.curr_balance += total
	wallet.save()
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)

	return Response({'status':1, 'message':'done'})

@api_view(['POST',])
@permission_classes((IsAuthenticated, ))
def ready_order(request):
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})

	products = stall.menu.all()
	try:
		sg_id = request.data['sg_id']
		stg = StallGroup.objects.get(id=sg_id)
		transaction = stg.transaction
		wallet = transaction.wallet
		stall = stg.stall
		if stg.cancelled:
			raise Exception
	except:
		return Response({'status':2, 'message':'Not Allowed'})
	stg.order_ready=True
	stg.save()
	db.child('stall').child(stall.id).child(stg.id).set(StallGroupSerializer(stg).data)
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	return Response({'status':1, 'message':'done'})


@api_view(['POST',])
@permission_classes((IsAuthenticated, ))
def open_close(request):
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	print stall
	# try:
	# 	request.POST['s']
	products = ProductMain.objects.filter(product__stall=stall)
	try:
		status = request.data['status']
		if status == 'open':
			products.update(is_available=True)
			message = 'succesfully opened'
		elif status == 'close':
			products.update(is_available=False)
			message = 'succesfully closed'
		return Response({'status':1, 'message':message})
	except:
		return Response({'status':2, 'message':'No status'})



@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass))
def order_complete(request):
	try:
		stall = request.user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	products = stall.menu.all()
	try:
		sg_id = request.data['sg_id']
		stg = StallGroup.objects.get(id=sg_id)
		transaction = stg.transaction
		wallet = transaction.wallet
		stall = stg.stall
		if not stg.order_ready:
			raise Exception
		if stg.cancelled:
			raise Exception
	except:
		return Response({'status':2, 'message':'Not Allowed'})
	stg.order_complete=True
	stg.save()
	db.child('stall').child(stall.id).child(stg.id).set(StallGroupSerializer(stg).data)
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	return Response({'status':1, 'message':'done'})


@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def generate_code(request):
	user = request.user
	data = request.data
	try:
		stall_group = StallGroup.objects.get(id=data['sg_id'])
		transaction = stall_group.transaction
		stall = stall_group.stall
		wallet = transaction.wallet
		if stall_group.cancelled:
			raise Exception
	except:
		return Response({'status':0, 'message':'not allowed'})
	if not stall_group.order_ready:
		return Response({'status':3, 'message':'Order not ready'})
	# stall_group.order_complete = True
	stall_group.code_requested = True
	stall_group.save()
	# db.child('stall').child(stall.id).child(stall_group.id).set(StallGroupSerializer(stall_group).data)
	# db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	return Response({'status':1})



#one time
def add_bitsian():
	from openpyxl import load_workbook
	path = os.path.join(BASE_DIR,'media', 'Bitsians.xlsx')
	wb = load_workbook(path)
	sheet = wb.get_sheet_by_name('Sheet')
	for i in range(101,4852):
		name = sheet.cell(row=i, column=2).value
		long_id = sheet.cell(row=i, column=1).value
		email = sheet.cell(row=i, column=6).value
		gender = sheet.cell(row=i, column=3).value
		room_no = int(sheet.cell(row=i, column=5).value)
		bhawan = sheet.cell(row=i, column=4).value
		Bitsian.objects.create(name=name, long_id=long_id, room_no=room_no,
			bhawan=bhawan,email=email, gender=gender)
	return;

#faltu
def get_attendance_count(prof_show):
	sum_ = 1
	for i in Attendance.objects.filter(prof_show=prof_show):
		sum_ += (i.count+i.passed_count)
	return sum_

############### 	ADMIN APP FOR DOLE AND AUDI FORCE  	###############

@api_view(['GET',])
@permission_classes((IsAuthenticated, ))
def get_profshows(request):
	user = request.user
	if user.is_superuser:
		prof_show_serializer = ProfShowSerializer(ProfShow.objects.all(), many=True)
		return Response(prof_show_serializer.data)
	clubdept = ClubDepartment.objects.get(user=user)
	prof_show_serializer = ProfShowSerializer(clubdept.profshows.all(), many=True)
	return Response(prof_show_serializer.data)

@api_view(['POST',])
@permission_classes((IsAuthenticated, ))
def validate_profshow(request):
	user = request.user
	data = request.data
	try:
		prof_show = ProfShow.objects.get(id=data['prof_show'])
	except:
		return Response({'status':0,'message':'Invalid Prof Show'})
	try:
		clubdepartment = ClubDepartment.objects.get(user=user)
		if not prof_show in clubdepartment.profshows.all():
			raise Exception
	except:
		return Response({'status':0, 'message':'Invalid Access'})
	try:
		uid = data['barcode']
		wallet = Wallet.objects.get(uid=uid)
	except:
		try:
			wallet = Wallet.objects.get(userid=uid)
		except:
			return Response({'status':0, 'message':'invalid code'})
	try:
		count = data['count']
	except:
		return Response({'status':0, 'message':'Invalid count'})
	is_bitsian = wallet.is_bitsian
	print count
	if is_bitsian:
		bitsian = wallet.bitsian
		bitsian_serializer = BitsianSerializer(bitsian)
		try:
			attendance = Attendance.objects.get(bitsian=bitsian, prof_show=prof_show)
		except:
			return Response({'status':0, 'message':'No more passes left. Please buy more passes.'})
		if attendance.count >= int(count):
			attendance.count -= int(count)
			attendance.passed_count += int(count)
			attendance.save()
			return Response({'status':1, 'message':'Success. Passes Left = ' + str(attendance.count), 'bitsian':bitsian_serializer.data})
		else:
			return Response({'status':0, 'message':'No more passes left. Please register at DoLE Booth.'})
		return Response({'status':0, 'message':'Please check barcode of Participant'})
	else:
		participant = wallet.participant
		participant_serializer = ParticipantSerializer(participant)
		try:
			attendance = Attendance.objects.get(participant=participant, prof_show=prof_show)
		except:
			return Response({'status':0, 'message':'No more passes left. Please register at DoLE Booth.'})
		if attendance.count >= int(count):
			attendance.count -= int(count)
			attendance.passed_count += int(count)
			attendance.save()
			return Response({'status':1, 'message':'Success. Passes Left = ' + str(attendance.count), 'participant':participant_serializer.data})
		else:
			return Response({'status':0, 'message':'No more passes left. Please buy more passes.'})






###########################	NO 		NEED	############################


@api_view(['GET','POST'])
@permission_classes((IsAuthenticated,))
def show_all_orders(request):
	user = request.user
	data = request.data
	try:
		stall = Stall.objects.get(user=user)
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	return Response({'Orders':StallGroupSerializer(StallGroup.objects.filter(stall=stall, order_complete=False, group_paid=True), many=True).data})

@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass))
def recieve_order(request):
	try:
		stall = user.stall
		if not stall:
			raise Exception
	except:
		return Response({'status':0, 'message':'Not Allowed'})
	products = stall.menu.all()
	try:
		wallet_id = request.data['id']
		wallet = Wallet.objects.get(uid = wallet_id)
		saleids = request.data.getlist('saleids')
		sales = Sale.objects.filter(uid__in=saleids, cancelled=False)
	except:
		return Response({'status':2, 'message':'invalid data'})
	sales.update(recieved=True, complete=True)
	return Response({'status':1})


@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def add_to_cart(request):
	data = request.data
	try:
		quantity = int(data['quantity'])
		if quantity <= 0:
			raise Exception
		p_id = data['p_id']
		product = ProductMain.objects.get(id=p_id)
		if not product.is_available:
			return Response({'status':2, 'message':product.__unicode__() + ' is not available right now.'})
	except:
		return Response({'status':0, 'message':'Invalid Input'})
	data = {
		'p_id': p_id,
		'price': product.price,
		'quantity': quantity,
		'name': product.product.name,
		}
	cart = Cart.objects.get(user=request.user)
	if quantity > product.quantity:
		return Response({'Not enough quantity.'})
	Sale.objects.create(cart=cart, product=product, quantity=quantity)
	cart.amount += (product.price*quantity)
	cart.save()
	cart_serializer = CartSerializer(cart)
	sale_serializer = SaleSerializer(Sale.objects.filter(cart=cart), many=True)
	return Response({'status':1, 'message':product.product.name + ' added succesfully to the cart', 'data':data, 'cart':cart_serializer.data, 'sales':sale_serializer.data})

@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass, UserPermissionClass))
def remove_from_cart(request):
	sale = Sale.objects.get(id=request.data['s_id'])
	product = sale.product
	quantity = sale.quantity
	sale.delete()
	cart = Cart.objects.get(user=request.user)
	cart.amount -= (product.price*quantity)
	cart_serializer = CartSerializer(cart)
	sale_serializer = SaleSerializer(Sale.objects.filter(cart=cart), many=True)
	return Response({'status':1, 'message':product.product.name + ' removed succesfully from the cart', 'cart':cart_serializer.data, 'sales':sale_serializer.data})

def add_bitsian():
	from openpyxl import load_workbook
	path = os.path.join(BASE_DIR,'media', 'Bitsians.xlsx')
	wb = load_workbook(path)
	sheet = wb.get_sheet_by_name('Sheet')
	for i in range(101,4852):
		name = sheet.cell(row=i, column=2).value
		long_id = sheet.cell(row=i, column=1).value
		email = sheet.cell(row=i, column=6).value
		gender = sheet.cell(row=i, column=3).value
		room_no = int(sheet.cell(row=i, column=5).value)
		bhawan = sheet.cell(row=i, column=4).value
		Bitsian.objects.create(name=name, long_id=long_id, room_no=room_no,
			bhawan=bhawan,email=email, gender=gender)
	return
@api_view(['POST',])
@permission_classes((IsAuthenticated, TokenVerificationClass, UserPermissionClass))
def checkout(request):
	try:
		bitsian = Bitsian.objects.get(user=request.user)
		sale_group = SaleGroup.objects.create(bitsian=bitsian, is_bitsian=True, user=request.user)
	except:
		try:
			participant = Participant.objects.get(user=request.user)
			sale_group = SaleGroup.objects.create(participant=participant, user=request.user)
		except:
			return Response({'status':0, 'message':'User does not exist'})
	try:
		wallet = request.user.wallet
		curr_balance = wallet.curr_balance
	except:
		return Response({'status':0, 'message':'Wallet does not exist'})

	for s_id in dict(request.data)['sale_ids']:
		sale = Sale.objects.get(id=int(s_id))
		if not sale.cart.user == request.user:
			sale_group.delete()
			return Response({'message':'Invalid access.'})
		sale.sale_group = sale_group
		sale_group.amount += (sale.product.price*sale.quantity)
		sale_group.save()
		sale.save()
	return Response({'id':sale_group.id, 'amount':sale_group.amount})


@api_view(['POST'])
@permission_classes((IsAuthenticated,))
def get_cart_details(request):
	try:
		stall = Stall.objects.get(user=request.user)
	except:
		return Response({'message':'Access denied'})

@api_view(['POST',])
@permission_classes((IsAuthenticated, UserPermissionClass, TokenVerificationClass))
def get_profile_bitsian(request):

	wallet_serializer = WalletSerializer(wallet)
	bitsian_serializer = BitsianSerializer(bitsian)
	profshow_serializer = AttendanceSerializer(Attendance.objects.filter(bitsian=bitsian), many=True)

	list_={}
	for t in wallet.transactions.all():
		list_[t.id] = TransactionSerializer(t).data
	dict_ = {'wallet':wallet_serializer.data, 'transactions': list_}
	db.child('wallet').child(wallet.uid).set(dict_)
	ts = TransactionSerializer(wallet.transactions.all(), many=True).data
	return Response({'bitsian':bitsian_serializer.data, 'prof_shows':profshow_serializer.data, 'wallet':wallet_serializer.data, 'username':username, 'password':password, 'transactions':ts})

@api_view(['POST'])
@permission_classes((IsAuthenticated, ))
def add_money_controls(request):
	user = request.user
	if not user.username == 'controls':
		return Response({'status':0, 'message':'Not Allowed'})
	try:
		uid = data['barcode']
		wallet = Wallet.objects.get(uid=uid)
	except:
		try:
			wallet = Wallet.objects.get(userid=uid)
		except:
			return Response({'status':0, 'message':'invalid code'})
	try:
		amount = int(data['count'])
		if amount <=0:
			raise Exception
	except:
		return Response({'status':0, 'message':'Invalid count'})

	wallet.curr_balance += amount
	wallet.save()
	transaction = Transaction.objects.create(wallet=wallet, value=amount, t_type="add", from_controls=True)
	db.child('wallet').child(wallet.uid).child('transactions').child(transaction.id).set(TransactionSerializer(transaction).data)
	db.child('wallet').child(wallet.uid).child('wallet').set(WalletSerializer(wallet).data)
	return Response({'status':1, 'message':str(amount) + ' Successfully added'})


def profshow_data():
	path = os.path.join(BASE_DIR, 'media', 'registered_lst.xlsx')
	prof_show = ProfShow.objects.get(id=2)
	from openpyxl import load_workbook
	wb = load_workbook(path)
	y=0
	for i in [['SR', 1, 403], ['RB',1 ,742 ], ['VKB',1 ,408 ], ['CVR',1 ,200 ], ['SV',1 ,456 ], ['Gandhi', 1, 380], ['MAL', 1, 575], ['Meera', 1, 640], ['RPA', 1, 298], ['DS', 1, 281]]:
		sheet = wb.get_sheet_by_name(i[0])
		for j in range(i[1], i[2]):
			try:
				x = sheet.cell(row=j, column=6).value
				if x is None:
					continue
				x = int(x)
				id_ = sheet.cell(row=j, column=1).value
				b = Bitsian.objects.get(long_id=id_)
				a = Attendance.objects.create(count=x, prof_show=prof_show, bitsian=b, number=get_attendance_count(prof_show))
				y+=1
			except Exception as e:
				print e
				print i, j
				# return 
	print y


def check_bitsian():
	path = os.path.join(BASE_DIR, 'media', 'bitsians1.xlsx')
	from openpyxl import load_workbook
	import re
	wb = load_workbook(path)
	sheet = wb.get_sheet_by_name('Sheet1')
	for i in range(2,4831):
		x = sheet.cell(row=i, column=1).value
		try:
			b = Bitsian.objects.get(long_id=x)
			continue
		except:
			year = x[:4]
			first = ''
			branch = x[4:8]
			if branch[:2] == 'C2':
				continue
			if branch[0] in ('A', 'B', 'D'):
				first = 'f'
			elif branch[0] in ('H', 'K'):
				first = 'h'
			elif branch[0] == 'P':
				first = 'p'
			else:
				continue
			id_ = x[8:]
			if not year == '2017':
				id_ = id_[1:]
			email = first + year + id_ + '@pilani.bits-pilani.ac.in'
			name = sheet.cell(row=i, column=2).value
			gender = sheet.cell(row=i, column=3).value
			bhawan = sheet.cell(row=i, column=4).value
			room_no = int(sheet.cell(row=i, column=5).value)
			b = Bitsian.objects.create(name=name, email=email, bhawan = bhawan, room_no=room_no, long_id=x)


def update_bitsians():
	for b in Bitsian.objects.all():
		
		x = b.long_id
		year = x[:4]
		first = ''
		branch = x[4:8]
		if branch[:2] == 'C2':
			continue
		if branch[0] in ('A', 'B', 'D'):
			first = 'f'
		elif branch[0] in ('H', 'K'):
			first = 'h'
		elif branch[0] == 'P':
			first = 'p'
		else:
			continue
		if not year == '2017':
			id_ = id_[1:]
		id_ = x[8:]
		email = first + year + id_ + '@pilani.bits-pilani.ac.in'
		b.email=email
		b.save()
		username = first + year + id_
		if b.user:
			try:
				user = b.user 
				user.username = username
				user.email = email
				user.save()
			except Exception as e:
				print e
				print user.username, user.email


def has_data():
	path = os.path.join(BASE_DIR, 'media', 'has.xlsx')
	prof_show = ProfShow.objects.get(id=3)
	from openpyxl import load_workbook
	wb = load_workbook(path)
	y=0
	z=0
	lists = [['DS', 1, 277, 4], ['MSA', 1, 140, 5], ['MB', 1, 685, 5], ['SR', 1, 402, 5], ['BD', 1, 363, 5],
	 ['RM', 1, 387, 5], ['ML-A', 1, 138,5 ], ['ML -B', 1, 133,5 ], ['GN', 1, 383,5 ], ['ML-C', 1, 171 ,5 ], ['VK', 1, 261 ,5 ],
	  ['BG', 1, 148 ,5 ], ['VY', 1, 250 ,5 ], ['AK', 1, 152 ,5 ], ['RP', 1, 148 ,5 ], ['CVR', 1, 200 ,5 ],
	   ['SK', 1, 208 ,5 ]]

	for i in lists:
		sheet = wb.get_sheet_by_name(i[0])
		for j in range(i[1], i[2]):
			try:
				x = sheet.cell(row=j, column=i[3]).value
				if x is None:
					continue
				x = int(x)
				id_ = sheet.cell(row=j, column=1).value
				z+=x
				b = Bitsian.objects.get(long_id=id_)
				try:
					a = Attendance.objects.create(count=x, prof_show=prof_show, bitsian=b, number=get_attendance_count(prof_show))
				except:
					pass
				y+=x
			except Exception as e:
				print e
				print i, j
	print y


def add_products():
	a=[('Schewan Chicken Noodles', '90'), ('Egg Biryani ', '80'), ('Chicken Biryani', '120'), ('Special Chicken Biryani', '160'), ('Egg Fried Maggi', '50'), ('Chicken fried Maggi', '70'), ('Chicken Burger', '60'), ('Chicken Cheese Burger', '70'), ('Plain Maggi', '25'), ('Veg Maggi', '30'), ('Fried Maggi', '35'), ('Cheese Maggi', '40'), ('Panneer Maggi', '50'), ('Cheese Fried maggi', '50'), ('Paneer Fried maggi', '55'), ('Veg Fried Rice', '45'), ('Schezwan Fried Rice', '55'), ('Paneer Fried Rice', '65'), ('Veg Hakka Noodles', '45'), ('Chilli Garlic Noodles', '55'), ('Schezwan Noodles', '55'), ('Paneer Fried Noodles', '65'), ('Veg Manchurian', '65'), ('Paneer Manchurian', '95'), ('Chilli Paneer', '95'), ('Chinese Combo', '95')]
	stall = Stall.objects.get(id=2)
	co = Colour.objects.get(id=1)
	pt = Type.objects.get(id=1)
	size = Size.objects.get(id=1)
	ct = PCategory.objects.get(id=1)
	for i in a:
		p1 = Product.objects.create(name=i[0], stall=stall, colour=co, p_type=pt, category=ct)
		p2 = ProductMain.objects.create(product=p1, size=size, price=int(i[1]))

def create_users():
	for b in Bitsian.objects.all():
		if not b.user:
			email = b.email
			password = ''.join(choice(chars) for _ in xrange(8))
			username = email.split('@')[0]
			try:
				user = User.objects.get(username=username)
				user.email = email
				user.set_password(password)
				user.save()
				b.user = user
				b.save()
			except:
				user = User.objects.create_user(username=username, password=password, email=email)
				b.user = user
				b.save()
